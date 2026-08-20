#!/usr/bin/env python3
import os
import re
import sys
import glob
from pathlib import Path

import xlrd
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

from dotenv import load_dotenv

load_dotenv()


DB_CONFIG = {
    'host': os.environ.get('KITCHEN_DB_HOST', 'localhost'),
    'database': os.environ.get('KITCHEN_DB_NAME', ''),
    'user': os.environ.get('KITCHEN_DB_USER', ''),
    'password': os.environ.get('KITCHEN_DB_PASSWORD', ''),
    'port': int(os.environ.get('KITCHEN_DB_PORT', '5432')),
}

BRAND_TRANSLIT = {
    "egger": "эггер", "rehau": "рехау", "boyard": "боярд", "hafele": "хефеле",
    "blum": "блюм", "kessebohmer": "кессебомер", "kronospan": "кроношпан",
    "aristo": "аристо", "rait": "райт", "gross": "гросс",
}

STOP_WORDS = {'ручка', 'петля', 'направляющая', 'доводчик'}


def _normalize(name):
    name = name.lower().strip().replace('х', 'x')
    for eng, rus in BRAND_TRANSLIT.items():
        name = name.replace(eng, rus)
    return re.sub(r'\s+', ' ', name).strip()


def _extract_search_key(name):
    words = [w for w in _normalize(name).split()
             if len(w) > 2 and w not in STOP_WORDS]
    return ' '.join(words[:4]) if words else name


def parse_materials(xls_path):
    wb = xlrd.open_workbook(xls_path)
    sheet = wb.sheet_by_index(0)

    header_row = -1
    for i in range(min(20, sheet.nrows)):
        row = [str(c).strip().lower() for c in sheet.row_values(i)]
        if 'наименование материала' in row:
            cols = {
                'name': row.index('наименование материала'),
                'qty': row.index('количество в изделии')
                    if 'количество в изделии' in row else 4,
                'unit': next((j for j, v in enumerate(row)
                              if 'ед. изм.' in v), 5),
            }
            header_row = i
            break

    if header_row == -1:
        print(f"  [ПРЕДУПРЕЖДЕНИЕ] Не найден заголовок 'Наименование материала'")
        return []

    materials = []
    for r in range(header_row + 1, sheet.nrows):
        vals = sheet.row_values(r)
        if len(vals) <= max(cols.values()):
            continue
        name = str(vals[cols['name']]).strip()
        if not name or name.lower() == 'итого':
            continue
        try:
            qty = float(vals[cols['qty']])
        except (ValueError, TypeError):
            qty = 0.0
        materials.append({
            'name': name,
            'qty': qty,
            'unit': str(vals[cols['unit']]).strip(),
        })

    return materials


def fetch_prices(product_names):
    if not product_names:
        return {}

    import psycopg2

    norm_map = {}
    for n in product_names:
        norm_map.setdefault(_normalize(n), []).append(n)

    res_prices = {}
    try:
        conn = psycopg2.connect(connect_timeout=5, **DB_CONFIG)
        cur = conn.cursor()

        norm_names = list(norm_map.keys())

        # Step 1: exact match
        cur.execute("""
            SELECT LOWER(TRIM(pt.name->>'ru_RU')),
                   COALESCE(psi.price, pt.list_price, 0.0)
            FROM product_template pt
            LEFT JOIN product_supplierinfo psi
                ON pt.id = psi.product_tmpl_id
            WHERE LOWER(TRIM(pt.name->>'ru_RU')) = ANY(%s)
        """, [norm_names])

        found_norms = set()
        for row in cur.fetchall():
            n_name, price = row[0], float(row[1])
            found_norms.add(n_name)
            for orig in norm_map[n_name]:
                res_prices[orig] = price

        # Step 2: pg_trgm similarity for missing
        missing = [n for n in norm_names if n not in found_norms]
        for m_norm in missing:
            if len(m_norm) < 5:
                continue
            try:
                cur.execute("""
                    SELECT
                        LOWER(TRIM(pt.name->>'ru_RU')) as db_name,
                        COALESCE(psi.price, pt.list_price, 0.0) as price,
                        similarity(pt.name->>'ru_RU', %s) as sim
                    FROM product_template pt
                    LEFT JOIN product_supplierinfo psi
                        ON pt.id = psi.product_tmpl_id
                    WHERE similarity(pt.name->>'ru_RU', %s) >= 0.4
                    ORDER BY sim DESC
                    LIMIT 1
                """, [m_norm, m_norm])
                row = cur.fetchone()
                if row:
                    price = float(row[1])
                    print(f"  [TRGM] '{m_norm}' ~ '{row[0]}' = {price}₽ "
                          f"(sim={row[2]:.3f})")
                    for orig in norm_map[m_norm]:
                        res_prices[orig] = price
                    found_norms.add(m_norm)
            except Exception:
                pass

        # Step 3: ILIKE fallback
        still_missing = [n for n in norm_names if n not in found_norms]
        for m_norm in still_missing:
            search = _extract_search_key(m_norm)
            if len(search) < 3:
                continue
            cur.execute("""
                SELECT COALESCE(psi.price, pt.list_price, 0.0)
                FROM product_template pt
                LEFT JOIN product_supplierinfo psi
                    ON pt.id = psi.product_tmpl_id
                WHERE pt.name->>'ru_RU' ILIKE %s
                LIMIT 1
            """, [f'%{search}%'])
            row = cur.fetchone()
            if row:
                print(f"  [ILIKE] '{m_norm}' -> '{search}' = {float(row[0]):.0f}₽")
                for orig in norm_map[m_norm]:
                    res_prices[orig] = float(row[0])

        cur.close()
        conn.close()
    except Exception as e:
        print(f"  [ОШИБКА БД] {e}", file=sys.stderr)
        return {n: 0.0 for n in product_names}

    return {n: res_prices.get(n, 0.0) for n in product_names}


_LDSP_PREFIXES = ('лдсп', 'дсп', 'хдф')
_LDSP_SHEET_M2 = 5.8

_WORKSHOP_KEYWORDS = ('лдсп', 'дсп', 'хдф', 'мдф', 'кромка')
_WORKSHOP_M2_RATE = 670
_SHEET_UNITS = {'м²', 'м2', 'кв.м', 'кв. м', 'm²', 'm2'}


def _is_ldsp(name):
    return name.lower().strip().startswith(_LDSP_PREFIXES)


def _is_workshop_material(name, unit):
    return name.lower().strip().startswith(_WORKSHOP_KEYWORDS) and unit.lower().replace(' ', '') in _SHEET_UNITS


def enrich_materials(materials):
    names = [m['name'] for m in materials]
    prices = fetch_prices(names)
    enriched = []
    for m in materials:
        cost = prices.get(m['name'], 0.0)
        if cost and _is_ldsp(m['name']):
            cost /= _LDSP_SHEET_M2
        enriched.append({**m, 'cost': cost})
    return enriched


def save_xlsx(materials, output_path, source_name, workshop_m2=0.0, workshop_cost=0.0):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Смета"

    header_font = Font(name='Arial', bold=True, size=11, color='FFFFFF')
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4',
                               fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center')
    cell_font = Font(name='Arial', size=10)
    cell_align = Alignment(vertical='center')
    money_fmt = '#,##0.00'
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin'),
    )

    headers = ['Наименование', 'Количество', 'Ед. изм.', 'Цена, ₽', 'Сумма, ₽']
    col_widths = [60, 12, 10, 14, 16]

    for col_idx, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    total_sum = 0.0
    found_count = 0
    total_count = len(materials)

    for row_idx, m in enumerate(materials, 2):
        name = m['name']
        qty = m['qty']
        unit = m.get('unit', '')
        cost = m.get('cost', 0.0)
        subtotal = cost * qty
        total_sum += subtotal
        if cost > 0:
            found_count += 1

        ws.cell(row=row_idx, column=1, value=name).font = cell_font
        ws.cell(row=row_idx, column=1).alignment = cell_align
        ws.cell(row=row_idx, column=1).border = thin_border

        ws.cell(row=row_idx, column=2, value=qty).font = cell_font
        ws.cell(row=row_idx, column=2).alignment = Alignment(horizontal='center', vertical='center')
        ws.cell(row=row_idx, column=2).border = thin_border

        ws.cell(row=row_idx, column=3, value=unit).font = cell_font
        ws.cell(row=row_idx, column=3).alignment = Alignment(horizontal='center', vertical='center')
        ws.cell(row=row_idx, column=3).border = thin_border

        price_cell = ws.cell(row=row_idx, column=4, value=cost)
        price_cell.font = cell_font
        price_cell.alignment = Alignment(horizontal='right', vertical='center')
        price_cell.number_format = money_fmt
        price_cell.border = thin_border

        sub_cell = ws.cell(row=row_idx, column=5, value=subtotal)
        sub_cell.font = cell_font
        sub_cell.alignment = Alignment(horizontal='right', vertical='center')
        sub_cell.number_format = money_fmt
        sub_cell.border = thin_border

    # итоговая строка (материалы)
    row = len(materials) + 2
    ws.cell(row=row, column=1, value=f"Итого материалы ({found_count}/{total_count})")
    ws.cell(row=row, column=1).font = Font(name='Arial', bold=True, size=11)
    ws.cell(row=row, column=1).border = thin_border
    for col in range(2, 5):
        ws.cell(row=row, column=col).border = thin_border
    total_cell = ws.cell(row=row, column=5, value=total_sum)
    total_cell.font = Font(name='Arial', bold=True, size=11)
    total_cell.alignment = Alignment(horizontal='right', vertical='center')
    total_cell.number_format = money_fmt
    total_cell.border = thin_border

    blank_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')

    # строка работы цеха
    if workshop_cost:
        row += 1
        ws.cell(row=row, column=1, value=f"Работа цеха ({workshop_m2:.2f} м² × {_WORKSHOP_M2_RATE} ₽)")
        ws.cell(row=row, column=1).font = Font(name='Arial', size=10)
        ws.cell(row=row, column=1).fill = blank_fill
        for col in range(2, 5):
            ws.cell(row=row, column=col).fill = blank_fill
        ws.cell(row=row, column=5, value=workshop_cost).font = Font(name='Arial', size=10)
        ws.cell(row=row, column=5).alignment = Alignment(horizontal='right', vertical='center')
        ws.cell(row=row, column=5).number_format = money_fmt
        ws.cell(row=row, column=5).fill = blank_fill

    # итого с работой цеха
    row += 1
    grand_total = total_sum + workshop_cost
    ws.cell(row=row, column=1, value="ИТОГО с работой цеха")
    ws.cell(row=row, column=1).font = Font(name='Arial', bold=True, size=12, color='C00000')
    ws.cell(row=row, column=1).border = thin_border
    for col in range(2, 5):
        ws.cell(row=row, column=col).border = thin_border
    gt_cell = ws.cell(row=row, column=5, value=grand_total)
    gt_cell.font = Font(name='Arial', bold=True, size=12, color='C00000')
    gt_cell.alignment = Alignment(horizontal='right', vertical='center')
    gt_cell.number_format = money_fmt
    gt_cell.border = thin_border

    wb.save(output_path)
    return total_sum, grand_total


def main():
    script_dir = Path(__file__).parent
    xls_dir = script_dir / 'xls'
    xls_dir.mkdir(parents=True, exist_ok=True)

    xls_files = sorted(glob.glob(str(xls_dir / '*.xls'))) + \
                sorted(glob.glob(str(xls_dir / '*.xlsx')))

    if not xls_files:
        print("Нет .xls/.xlsx файлов в папке 'xls/'.")
        print(f"Положите файл со сметой в: {xls_dir}")
        return

    for xls_path in xls_files:
        xls_path = Path(xls_path)
        print(f"\n{'='*60}")
        print(f"Файл: {xls_path.name}")
        print(f"{'='*60}")

        materials = parse_materials(str(xls_path))
        if not materials:
            print("  Не удалось распарсить материалы.")
            continue

        print(f"  Найдено позиций: {len(materials)}")
        print("  Загружаем цены из БД...")

        enriched = enrich_materials(materials)

        workshop_m2 = sum(m['qty'] for m in enriched if _is_workshop_material(m['name'], m.get('unit', '')))
        workshop_cost = workshop_m2 * _WORKSHOP_M2_RATE

        output_name = xls_path.stem + '_с_ценами.xlsx'
        output_path = script_dir / output_name
        total, grand_total = save_xlsx(
            enriched, str(output_path), xls_path.name,
            workshop_m2=workshop_m2, workshop_cost=workshop_cost,
        )

        found = sum(1 for m in enriched if m.get('cost', 0) > 0)
        not_found = [m['name'] for m in enriched if m.get('cost', 0) <= 0]
        print(f"  Найдено цен: {found}/{len(enriched)}")
        print(f"  Материалы: {total:,.2f} ₽".replace(',', ' '))
        print(f"  Работа цеха ({workshop_m2:.2f} м²): {workshop_cost:,.2f} ₽".replace(',', ' '))
        print(f"  Итого с работой цеха: {grand_total:,.2f} ₽".replace(',', ' '))
        if not_found:
            print(f"  Не найдены цены для:")
            for n in not_found:
                print(f"    • {n}")
        print(f"  Результат: {output_path}")


if __name__ == '__main__':
    main()
